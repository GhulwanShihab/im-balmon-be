"""
Device Export Service - Laporan Statistik Penggunaan Perangkat
"""
from typing import List, Optional, Dict, Any
from datetime import datetime, date
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, or_, extract
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import logging

from src.models.perangkat import Device, DeviceStatus
from src.models.device_child import DeviceChild
from src.models.loan import DeviceLoan, DeviceLoanItem, LoanStatus

logger = logging.getLogger(__name__)

NAMA_BULAN = [
    "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember"
]

def _format_tanggal(dt) -> str:
    """Format datetime ke 'dd MMMM yyyy' locale Indonesia"""
    if not dt:
        return "-"
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt)
        except ValueError:
            return dt
    day = dt.day
    month = NAMA_BULAN[dt.month] if 1 <= dt.month <= 12 else str(dt.month)
    year = dt.year
    return f"{day} {month} {year}"

def _format_tanggal_waktu(dt) -> str:
    """Format datetime ke 'dd MMMM yyyy HH:MM' locale Indonesia"""
    if not dt:
        return "-"
    day = dt.day
    month = NAMA_BULAN[dt.month] if 1 <= dt.month <= 12 else str(dt.month)
    year = dt.year
    return f"{day} {month} {year} {dt.strftime('%H:%M')}"


class DeviceExportService:
    def __init__(self, session: AsyncSession):
        self.session = session

    async def export_device_usage_to_excel(
        self,
        year: Optional[int] = None,
        month: Optional[int] = None,
        device_ids: Optional[List[int]] = None
    ) -> BytesIO:
        """
        Export statistik penggunaan perangkat ke Excel
        """
        logger.info("Starting device usage export...")
        
        try:
            wb = Workbook()
            
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            devices_data = await self._get_devices_with_usage(year, month, device_ids)
            monthly_stats = await self._get_monthly_stats(year, device_ids)
            yearly_stats = await self._get_yearly_stats(device_ids)
            usage_details = await self._get_usage_details(year, month, device_ids)
            
            # Buat semua sheet
            self._create_dashboard_sheet(wb, devices_data, monthly_stats, yearly_stats, year, month)
            self._create_device_summary_sheet(wb, devices_data)
            self._create_monthly_stats_sheet(wb, monthly_stats)
            self._create_yearly_stats_sheet(wb, yearly_stats)
            self._create_usage_details_sheet(wb, usage_details)
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            logger.info(f"Export completed: {buffer.getbuffer().nbytes} bytes")
            return buffer
            
        except Exception as e:
            logger.error(f"Error in export: {str(e)}", exc_info=True)
            raise

    async def _get_devices_with_usage(
        self,
        year: Optional[int] = None,
        month: Optional[int] = None,
        device_ids: Optional[List[int]] = None
    ) -> List[Dict[str, Any]]:
        """Get devices with usage statistics"""
        try:
            query = select(Device)
            if device_ids:
                query = query.where(Device.id.in_(device_ids))
            
            result = await self.session.execute(query)
            devices = result.scalars().all()
            
            devices_data = []
            for device in devices:
                # Get loan statistics for this device (parent level)
                loan_query = select(DeviceLoanItem).join(DeviceLoan).where(
                    and_(
                        DeviceLoanItem.device_id == device.id,
                        DeviceLoan.status.in_(['ACTIVE', 'RETURNED', 'OVERDUE'])
                    )
                )
                
                if year:
                    loan_query = loan_query.where(extract('year', DeviceLoan.loan_start_date) == year)
                if month:
                    loan_query = loan_query.where(extract('month', DeviceLoan.loan_start_date) == month)
                
                loan_result = await self.session.execute(loan_query)
                loan_items = loan_result.scalars().all()
                
                # Calculate statistics
                total_loans = len(set(item.loan_id for item in loan_items))
                total_usage_days = 0
                first_used_date = None
                last_used_date = None
                
                for item in loan_items:
                    loan = await self.session.get(DeviceLoan, item.loan_id)
                    if loan and loan.usage_duration_days:
                        total_usage_days += loan.usage_duration_days
                        
                        if first_used_date is None or loan.loan_start_date < first_used_date:
                            first_used_date = loan.loan_start_date
                        
                        if last_used_date is None or (loan.loan_end_date and loan.loan_end_date > last_used_date):
                            last_used_date = loan.loan_end_date
                
                # Get children count and their usage
                children_query = select(DeviceChild).where(DeviceChild.parent_id == device.id)
                children_result = await self.session.execute(children_query)
                children = children_result.scalars().all()
                
                total_children = len(children)
                children_available = sum(1 for c in children if c.device_status == DeviceStatus.TERSEDIA)
                children_borrowed = sum(1 for c in children if c.device_status == DeviceStatus.DIPINJAM)
                
                # Get child devices usage
                children_data = []
                for child in children:
                    child_loan_query = select(DeviceLoanItem).join(DeviceLoan).where(
                        and_(
                            DeviceLoanItem.child_device_id == child.id,
                            DeviceLoan.status.in_(['ACTIVE', 'RETURNED', 'OVERDUE'])
                        )
                    )
                    
                    if year:
                        child_loan_query = child_loan_query.where(extract('year', DeviceLoan.loan_start_date) == year)
                    if month:
                        child_loan_query = child_loan_query.where(extract('month', DeviceLoan.loan_start_date) == month)
                    
                    child_loan_result = await self.session.execute(child_loan_query)
                    child_loan_items = child_loan_result.scalars().all()
                    
                    child_total_loans = len(set(item.loan_id for item in child_loan_items))
                    child_usage_days = 0
                    
                    for item in child_loan_items:
                        loan = await self.session.get(DeviceLoan, item.loan_id)
                        if loan and loan.usage_duration_days:
                            child_usage_days += loan.usage_duration_days
                    
                    children_data.append({
                        "device_name": child.device_name,
                        "device_code": child.device_code,
                        "nup_device": child.nup_device,
                        "device_brand": child.bmn_brand or child.sample_brand or "-",
                        "device_condition": child.device_condition or "-",
                        "device_status": child.device_status.value if child.device_status else "TERSEDIA",
                        "total_loans": child_total_loans,
                        "total_usage_days": child_usage_days,
                    })
                
                devices_data.append({
                    "device_id": device.id,
                    "device_name": device.device_name,
                    "device_code": device.device_code,
                    "nup_device": device.nup_device,
                    "device_brand": device.bmn_brand or device.sample_brand or "-",
                    "device_year": device.device_year,
                    "device_type": device.device_type or "-",
                    "device_station": device.device_station or "-",
                    "device_room": device.device_room or "-",
                    "device_condition": device.device_condition or "-",
                    "device_status": device.device_status.value if device.device_status else "TERSEDIA",
                    "created_at": device.created_at,
                    "total_loans": total_loans,
                    "total_usage_days": total_usage_days,
                    "first_used_date": first_used_date,
                    "last_used_date": last_used_date,
                    "total_children": total_children,
                    "children_available": children_available,
                    "children_borrowed": children_borrowed,
                    "children": children_data,
                })
            
            return devices_data
            
        except Exception as e:
            logger.error(f"Error getting devices with usage: {str(e)}", exc_info=True)
            raise

    async def _get_monthly_stats(self, year: Optional[int], device_ids: Optional[List[int]]) -> List[Dict]:
        """Get monthly statistics"""
        try:
            query = select(
                extract('year', DeviceLoan.loan_start_date).label('year'),
                extract('month', DeviceLoan.loan_start_date).label('month'),
                func.count(func.distinct(DeviceLoan.id)).label('total_loans'),
                func.count(func.distinct(DeviceLoanItem.device_id)).label('unique_devices'),
                func.avg(DeviceLoan.usage_duration_days).label('avg_duration')
            ).select_from(DeviceLoan).join(DeviceLoanItem).where(
                DeviceLoan.status.in_(['ACTIVE', 'RETURNED', 'OVERDUE'])
            ).group_by(
                extract('year', DeviceLoan.loan_start_date),
                extract('month', DeviceLoan.loan_start_date)
            ).order_by(
                extract('year', DeviceLoan.loan_start_date).desc(),
                extract('month', DeviceLoan.loan_start_date).desc()
            )
            
            if year:
                query = query.where(extract('year', DeviceLoan.loan_start_date) == year)
            
            if device_ids:
                query = query.where(DeviceLoanItem.device_id.in_(device_ids))
            
            result = await self.session.execute(query)
            rows = result.all()
            
            monthly_stats = []
            for row in rows:
                monthly_stats.append({
                    "year": int(row.year) if row.year else 0,
                    "month": int(row.month) if row.month else 0,
                    "total_loans": row.total_loans or 0,
                    "unique_devices": row.unique_devices or 0,
                    "avg_duration": float(row.avg_duration) if row.avg_duration else 0,
                })
            
            return monthly_stats
            
        except Exception as e:
            logger.error(f"Error getting monthly stats: {str(e)}", exc_info=True)
            return []

    async def _get_yearly_stats(self, device_ids: Optional[List[int]]) -> List[Dict]:
        """Get yearly statistics"""
        try:
            query = select(
                extract('year', DeviceLoan.loan_start_date).label('year'),
                func.count(func.distinct(DeviceLoan.id)).label('total_loans'),
                func.count(func.distinct(DeviceLoanItem.device_id)).label('unique_devices'),
                func.avg(DeviceLoan.usage_duration_days).label('avg_duration')
            ).select_from(DeviceLoan).join(DeviceLoanItem).where(
                DeviceLoan.status.in_(['ACTIVE', 'RETURNED', 'OVERDUE'])
            ).group_by(
                extract('year', DeviceLoan.loan_start_date)
            ).order_by(
                extract('year', DeviceLoan.loan_start_date).desc()
            )
            
            if device_ids:
                query = query.where(DeviceLoanItem.device_id.in_(device_ids))
            
            result = await self.session.execute(query)
            rows = result.all()
            
            yearly_stats = []
            for row in rows:
                yearly_stats.append({
                    "year": int(row.year) if row.year else 0,
                    "total_loans": row.total_loans or 0,
                    "unique_devices": row.unique_devices or 0,
                    "avg_duration": float(row.avg_duration) if row.avg_duration else 0,
                })
            
            return yearly_stats
            
        except Exception as e:
            logger.error(f"Error getting yearly stats: {str(e)}", exc_info=True)
            return []

    async def _get_usage_details(
        self, 
        year: Optional[int], 
        month: Optional[int], 
        device_ids: Optional[List[int]]
    ) -> List[Dict]:
        """Get detailed usage records"""
        try:
            query = select(DeviceLoanItem).join(DeviceLoan).where(
                DeviceLoan.status.in_(['ACTIVE', 'RETURNED', 'OVERDUE'])
            ).order_by(DeviceLoan.loan_start_date.desc()).limit(100)
            
            if year:
                query = query.where(extract('year', DeviceLoan.loan_start_date) == year)
            if month:
                query = query.where(extract('month', DeviceLoan.loan_start_date) == month)
            if device_ids:
                query = query.where(DeviceLoanItem.device_id.in_(device_ids))
            
            result = await self.session.execute(query)
            loan_items = result.scalars().all()
            
            usage_records = []
            for item in loan_items:
                loan = await self.session.get(DeviceLoan, item.loan_id)
                device = await self.session.get(Device, item.device_id) if item.device_id else None
                child = await self.session.get(DeviceChild, item.child_device_id) if item.child_device_id else None
                
                if loan:
                    # Translate status
                    status_map = {
                        'ACTIVE': 'Aktif',
                        'RETURNED': 'Dikembalikan',
                        'OVERDUE': 'Terlambat',
                        'CANCELLED': 'Dibatalkan',
                    }
                    
                    usage_records.append({
                        "loan_number": loan.loan_number,
                        "borrower_name": loan.borrower_name,
                        "activity_name": loan.activity_name,
                        "loan_start_date": loan.loan_start_date,
                        "loan_end_date": loan.loan_end_date,
                        "usage_duration_days": loan.usage_duration_days,
                        "status": status_map.get(loan.status, loan.status),
                        "device_name": device.device_name if device else (child.device_name if child else "-"),
                        "device_code": device.device_code if device else (child.device_code if child else "-"),
                        "quantity": item.quantity,
                    })
            
            return usage_records
            
        except Exception as e:
            logger.error(f"Error getting usage details: {str(e)}", exc_info=True)
            return []

    def _create_device_summary_sheet(self, wb: Workbook, devices_data: List[Dict[str, Any]]):
        """Create device summary sheet — Ringkasan Perangkat"""
        try:
            ws = wb.create_sheet("Ringkasan Perangkat")
            
            headers = [
                "No", "Nama Perangkat", "Kode Perangkat", "NUP", "Merek", 
                "Tahun", "Tipe", "Stasiun", "Ruangan", "Kondisi", "Status",
                "Jumlah Peminjaman", "Lama Pakai (Hari)",
                "Pertama Dipakai", "Terakhir Dipakai",
                "Jml Anak", "Tersedia", "Dipinjam"
            ]
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Translate condition and status
            kondisi_map = {"BAIK": "Baik", "RUSAK": "Rusak", "MAINTENANCE": "Maintenance"}
            status_map = {"TERSEDIA": "Tersedia", "DIPINJAM": "Dipinjam", "MAINTENANCE": "Maintenance", "NONAKTIF": "Tidak Aktif"}
            
            row_num = 2
            for idx, device in enumerate(devices_data, 1):
                ws.cell(row=row_num, column=1).value = idx
                ws.cell(row=row_num, column=2).value = device["device_name"]
                ws.cell(row=row_num, column=3).value = device["device_code"]
                ws.cell(row=row_num, column=4).value = device["nup_device"]
                ws.cell(row=row_num, column=5).value = device["device_brand"]
                ws.cell(row=row_num, column=6).value = device["device_year"]
                ws.cell(row=row_num, column=7).value = device["device_type"]
                ws.cell(row=row_num, column=8).value = device["device_station"]
                ws.cell(row=row_num, column=9).value = device["device_room"]
                ws.cell(row=row_num, column=10).value = kondisi_map.get(device["device_condition"].upper(), device["device_condition"]) if device["device_condition"] != "-" else "-"
                ws.cell(row=row_num, column=11).value = status_map.get(device["device_status"].upper(), device["device_status"]) if device["device_status"] else "-"
                ws.cell(row=row_num, column=12).value = device["total_loans"]
                ws.cell(row=row_num, column=13).value = device["total_usage_days"]
                ws.cell(row=row_num, column=14).value = _format_tanggal(device["first_used_date"])
                ws.cell(row=row_num, column=15).value = _format_tanggal(device["last_used_date"])
                ws.cell(row=row_num, column=16).value = device["total_children"]
                ws.cell(row=row_num, column=17).value = device["children_available"]
                ws.cell(row=row_num, column=18).value = device["children_borrowed"]
                
                # Add border to all cells in this row
                for col in range(1, 19):
                    ws.cell(row=row_num, column=col).border = thin_border
                
                row_num += 1
                
                # Add child devices
                if device["children"]:
                    child_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    for child_idx, child in enumerate(device["children"], 1):
                        ws.cell(row=row_num, column=1).value = f"  {idx}.{child_idx}"
                        ws.cell(row=row_num, column=2).value = f"  └─ {child['device_name']}"
                        ws.cell(row=row_num, column=3).value = child['device_code']
                        ws.cell(row=row_num, column=4).value = child['nup_device']
                        ws.cell(row=row_num, column=5).value = child['device_brand']
                        ws.cell(row=row_num, column=10).value = kondisi_map.get(child['device_condition'].upper(), child['device_condition']) if child['device_condition'] != "-" else "-"
                        ws.cell(row=row_num, column=11).value = status_map.get(child['device_status'].upper(), child['device_status']) if child['device_status'] else "-"
                        ws.cell(row=row_num, column=12).value = child['total_loans']
                        ws.cell(row=row_num, column=13).value = child['total_usage_days']
                        
                        for col in range(1, 19):
                            ws.cell(row=row_num, column=col).fill = child_fill
                            ws.cell(row=row_num, column=col).border = thin_border
                        
                        row_num += 1
            
            # Adjust column widths
            column_widths = [6, 30, 15, 12, 20, 8, 15, 15, 15, 12, 12, 15, 14, 18, 18, 10, 10, 10]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col_num)].width = width
            
            ws.freeze_panes = "A2"
            
        except Exception as e:
            logger.error(f"Error creating device summary sheet: {str(e)}", exc_info=True)
            raise

    def _create_monthly_stats_sheet(self, wb: Workbook, monthly_stats: List[Dict]):
        """Create monthly statistics sheet — Statistik Bulanan"""
        try:
            ws = wb.create_sheet("Statistik Bulanan")
            
            headers = ["Tahun", "Bulan", "Nama Bulan", "Jumlah Peminjaman", "Perangkat Unik", "Rata-rata Durasi (Hari)"]
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            for row_num, stat in enumerate(monthly_stats, 2):
                ws.cell(row=row_num, column=1).value = stat["year"]
                ws.cell(row=row_num, column=2).value = stat["month"]
                ws.cell(row=row_num, column=3).value = NAMA_BULAN[stat["month"]] if 1 <= stat["month"] <= 12 else ""
                ws.cell(row=row_num, column=4).value = stat["total_loans"]
                ws.cell(row=row_num, column=5).value = stat["unique_devices"]
                ws.cell(row=row_num, column=6).value = round(stat["avg_duration"], 2)
                
                for col in range(1, 7):
                    ws.cell(row=row_num, column=col).border = thin_border
            
            for col in range(1, 7):
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            ws.freeze_panes = "A2"
            
        except Exception as e:
            logger.error(f"Error creating monthly stats sheet: {str(e)}", exc_info=True)

    def _create_yearly_stats_sheet(self, wb: Workbook, yearly_stats: List[Dict]):
        """Create yearly statistics sheet — Statistik Tahunan"""
        try:
            ws = wb.create_sheet("Statistik Tahunan")
            
            headers = ["Tahun", "Jumlah Peminjaman", "Perangkat Unik", "Rata-rata Durasi (Hari)"]
            header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            header_font = Font(bold=True, color="000000", size=10)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            for row_num, stat in enumerate(yearly_stats, 2):
                ws.cell(row=row_num, column=1).value = stat["year"]
                ws.cell(row=row_num, column=2).value = stat["total_loans"]
                ws.cell(row=row_num, column=3).value = stat["unique_devices"]
                ws.cell(row=row_num, column=4).value = round(stat["avg_duration"], 2)
                
                for col in range(1, 5):
                    ws.cell(row=row_num, column=col).border = thin_border
            
            for col in range(1, 5):
                ws.column_dimensions[get_column_letter(col)].width = 22
            
            ws.freeze_panes = "A2"
            
        except Exception as e:
            logger.error(f"Error creating yearly stats sheet: {str(e)}", exc_info=True)

    def _create_usage_details_sheet(self, wb: Workbook, usage_records: List[Dict]):
        """Create detailed usage records sheet — Detail Penggunaan"""
        try:
            ws = wb.create_sheet("Detail Penggunaan")
            
            headers = [
                "No. Peminjaman", "Nama Peminjam", "Nama Kegiatan", 
                "Tanggal Mulai", "Tanggal Selesai",
                "Durasi (Hari)", "Status", "Nama Perangkat", "Kode Perangkat", "Jumlah"
            ]
            header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            for row_num, record in enumerate(usage_records, 2):
                ws.cell(row=row_num, column=1).value = record["loan_number"]
                ws.cell(row=row_num, column=2).value = record["borrower_name"]
                ws.cell(row=row_num, column=3).value = record["activity_name"]
                ws.cell(row=row_num, column=4).value = _format_tanggal(record["loan_start_date"])
                ws.cell(row=row_num, column=5).value = _format_tanggal(record["loan_end_date"])
                ws.cell(row=row_num, column=6).value = record["usage_duration_days"]
                ws.cell(row=row_num, column=7).value = record["status"]
                ws.cell(row=row_num, column=8).value = record["device_name"]
                ws.cell(row=row_num, column=9).value = record["device_code"]
                ws.cell(row=row_num, column=10).value = record["quantity"]
                
                for col in range(1, 11):
                    ws.cell(row=row_num, column=col).border = thin_border
            
            column_widths = [18, 25, 30, 18, 18, 14, 14, 30, 15, 10]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col_num)].width = width
            
            ws.freeze_panes = "A2"
            
        except Exception as e:
            logger.error(f"Error creating usage details sheet: {str(e)}", exc_info=True)

    def _create_dashboard_sheet(
        self, 
        wb: Workbook, 
        devices_data: List[Dict], 
        monthly_stats: List[Dict],
        yearly_stats: List[Dict],
        year: Optional[int] = None,
        month: Optional[int] = None
    ):
        """Create dashboard summary sheet"""
        try:
            ws = wb.create_sheet("Dashboard", 0)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells("A1:F1")
            title_cell = ws["A1"]
            title_cell.value = "LAPORAN STATISTIK PENGGUNAAN PERANGKAT"
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 35
            
            # Subtitle with filter info
            ws.merge_cells("A2:F2")
            subtitle = "Balai Monitor Spektrum Frekuensi Radio"
            ws["A2"] = subtitle
            ws["A2"].font = Font(size=10, color="666666", italic=True)
            ws["A2"].alignment = Alignment(horizontal="center")
            
            # Report info
            ws["A4"] = "Tanggal Laporan:"
            ws["B4"] = _format_tanggal_waktu(datetime.now())
            ws["A4"].font = Font(bold=True, size=10)
            ws["B4"].font = Font(size=10)
            
            # Filter info
            if year or month:
                ws["A5"] = "Periode Filter:"
                period_parts = []
                if month and 1 <= month <= 12:
                    period_parts.append(NAMA_BULAN[month])
                if year:
                    period_parts.append(str(year))
                ws["B5"] = " ".join(period_parts) if period_parts else "Semua Periode"
                ws["A5"].font = Font(bold=True, size=10)
                ws["B5"].font = Font(size=10)
            
            # Summary section
            ws.merge_cells("A7:F7")
            ws["A7"] = "RINGKASAN KESELURUHAN"
            ws["A7"].font = Font(bold=True, size=12)
            ws["A7"].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            total_devices = len(devices_data)
            devices_with_usage = sum(1 for d in devices_data if d["total_loans"] > 0)
            total_loans = sum(d["total_loans"] for d in devices_data)
            total_usage_days = sum(d["total_usage_days"] for d in devices_data)
            
            summary_data = [
                ("Total Perangkat", total_devices),
                ("Perangkat Pernah Dipakai", devices_with_usage),
                ("Total Peminjaman", total_loans),
                ("Total Lama Pakai (Hari)", total_usage_days),
                ("Rata-rata Pakai per Perangkat (Hari)", round(total_usage_days / total_devices, 2) if total_devices > 0 else 0),
            ]
            
            row = 9
            for label, value in summary_data:
                ws.cell(row=row, column=1).value = label
                ws.cell(row=row, column=1).font = Font(bold=True, size=10)
                ws.cell(row=row, column=2).value = value
                ws.cell(row=row, column=2).font = Font(size=10)
                ws.cell(row=row, column=1).border = thin_border
                ws.cell(row=row, column=2).border = thin_border
                row += 1
            
            # Top devices section
            row += 1
            ws.merge_cells(f"A{row}:F{row}")
            ws.cell(row=row, column=1).value = "TOP 10 PERANGKAT PALING SERING DIPAKAI"
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            row += 2
            top_headers = ["Peringkat", "Nama Perangkat", "NUP", "Jumlah Peminjaman", "Lama Pakai (Hari)"]
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            for col_num, header in enumerate(top_headers, 1):
                cell = ws.cell(row=row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            
            sorted_devices = sorted(devices_data, key=lambda x: x["total_usage_days"], reverse=True)[:10]
            row += 1
            for rank, device in enumerate(sorted_devices, 1):
                ws.cell(row=row, column=1).value = rank
                ws.cell(row=row, column=2).value = device["device_name"]
                ws.cell(row=row, column=3).value = device["nup_device"]
                ws.cell(row=row, column=4).value = device["total_loans"]
                ws.cell(row=row, column=5).value = device["total_usage_days"]
                
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = thin_border
                
                row += 1
            
            # Adjust widths
            ws.column_dimensions["A"].width = 35
            ws.column_dimensions["B"].width = 35
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 20
            ws.column_dimensions["E"].width = 18
            ws.column_dimensions["F"].width = 15
            
        except Exception as e:
            logger.error(f"Error creating dashboard sheet: {str(e)}", exc_info=True)